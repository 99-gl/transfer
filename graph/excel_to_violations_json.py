"""将违例 Excel 表转换为 Graphiti 导入所需的 JSON。

Excel 从第 3 行起读取，前四列依次为：序号、违例概念、现象、识别方法。
序号、违例概念和识别方法允许使用纵向合并单元格。
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any


def cell_text(value: Any) -> str:
    """将 Excel 单元格内容标准化为去除首尾空白的字符串。"""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def make_id(prefix: str, serial: str, row_number: int) -> str:
    """使用表中序号生成示例风格的 ID；空序号时以 Excel 行号兜底。"""
    try:
        number = int(serial)
    except ValueError:
        suffix = serial or str(row_number - 2)
        return f"{prefix}_{suffix}"
    return f"{prefix}_{number:03d}"


def convert_excel(input_path: Path, output_path: Path, sheet_name: str | None) -> tuple[int, int]:
    """读取工作表并写入 ``nodes`` / ``edges`` JSON，返回节点数和边数。"""
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("缺少依赖 openpyxl，请先执行：python -m pip install openpyxl") from exc

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        if sheet_name is None:
            worksheet = workbook.active
        elif sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            raise ValueError(f"未找到工作表 {sheet_name!r}；可用工作表：{', '.join(workbook.sheetnames)}")

        nodes: list[dict[str, Any]] = []
        edges: list[dict[str, str]] = []
        violation_ids: set[str] = set()
        phenomenon_index = 0
        current_violation_id = ""
        current_identification_method = ""

        for row_number, row in enumerate(worksheet.iter_rows(min_row=3, max_col=4, values_only=True), start=3):
            serial, violation_name, phenomenon_name, identification_method = map(cell_text, row)

            # 完全空白的行不产生任何数据。纵向合并单元格只有首个单元格有值，
            # 所以序号、违例概念和识别方法都要在所属分组内向下继承。
            if not any((serial, violation_name, phenomenon_name, identification_method)):
                continue

            if bool(serial) != bool(violation_name):
                raise ValueError(f"第 {row_number} 行的序号和违例概念必须同时填写，或同时留空")
            if serial:
                # 识别方法不能跨违例分组继承。
                current_identification_method = ""
                current_violation_id = make_id("v", serial, row_number)
                if current_violation_id in violation_ids:
                    raise ValueError(f"第 {row_number} 行序号 {serial!r} 重复，无法生成唯一违例节点 ID")
                violation_ids.add(current_violation_id)
                nodes.append(
                    {
                        "id": current_violation_id,
                        "type": "ViolationConcept",
                        "properties": {"name": violation_name},
                    }
                )
            elif not current_violation_id:
                raise ValueError(f"第 {row_number} 行缺少序号和违例概念，且前面没有可继承的分组")

            if not phenomenon_name:
                raise ValueError(f"第 {row_number} 行缺少现象")
            if identification_method:
                current_identification_method = identification_method
            elif not current_identification_method:
                raise ValueError(f"第 {row_number} 行缺少识别方法，且前面没有可继承的值")

            phenomenon_index += 1
            phenomenon_id = f"p_{phenomenon_index:03d}"
            nodes.append(
                {
                    "id": phenomenon_id,
                    "type": "Phenomenon",
                    "properties": {
                        "name": phenomenon_name,
                        "identification_method": current_identification_method,
                    },
                }
            )
            edges.append(
                {"source": current_violation_id, "target": phenomenon_id, "relation": "has_phenomenon"}
            )
    finally:
        workbook.close()

    result = {"nodes": nodes, "edges": edges}
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as file:
        json.dump(result, file, ensure_ascii=False, indent=2)
        file.write("\n")
    return len(nodes), len(edges)


def main() -> None:
    parser = argparse.ArgumentParser(description="将违例 Excel（第 3 行起）转换为图谱 JSON")
    parser.add_argument("input", type=Path, help="输入 .xlsx 文件")
    parser.add_argument(
        "-o", "--output", type=Path, default=Path("violations_data.json"), help="输出 JSON 路径"
    )
    parser.add_argument("--sheet", help="工作表名称；默认读取活动工作表")
    args = parser.parse_args()

    if args.input.suffix.lower() not in {".xlsx", ".xlsm"}:
        parser.error("仅支持 .xlsx 或 .xlsm 文件")
    if not args.input.is_file():
        parser.error(f"输入文件不存在：{args.input}")

    try:
        node_count, edge_count = convert_excel(args.input, args.output, args.sheet)
    except (OSError, RuntimeError, ValueError) as exc:
        print(f"转换失败：{exc}", file=sys.stderr)
        raise SystemExit(1) from exc
    print(f"转换完成：{args.output}（{node_count} 个节点，{edge_count} 条边）")


if __name__ == "__main__":
    main()
