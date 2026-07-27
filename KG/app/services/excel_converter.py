"""Convert the current violation Excel template into a normalized graph snapshot."""

from __future__ import annotations

from io import BytesIO
from typing import Any
from uuid import UUID, uuid5

from app.schemas import GraphEdge, GraphNode, GraphSnapshot

UUID_NAMESPACE = UUID('d1cd9102-cd4b-49eb-92fa-046c813cc519')


def stable_uuid(kind: str, source_id: str) -> str:
    """Return an ID stable across imports while mutable properties can change."""
    return str(uuid5(UUID_NAMESPACE, f'{kind}:{source_id}'))


def _cell_text(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _load_workbook(content: bytes):
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError('缺少依赖 openpyxl，请安装应用依赖后重试。') from exc
    return load_workbook(BytesIO(content), read_only=True, data_only=True)


def convert_excel(content: bytes, sheet_name: str | None = None) -> GraphSnapshot:
    """Parse columns: serial, violation concept, phenomenon, identification method.

    Data begins at row 3. Merged serial/concept/method cells are inherited from
    the preceding non-empty row, matching the original conversion script.
    """
    workbook = _load_workbook(content)
    try:
        if sheet_name is None:
            worksheet = workbook.active
        elif sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            names = ', '.join(workbook.sheetnames)
            raise ValueError(f'未找到工作表 {sheet_name!r}；可用工作表：{names}')

        nodes: list[GraphNode] = []
        edges: list[GraphEdge] = []
        node_ids: set[str] = set()
        edge_ids: set[str] = set()
        current_violation_id = ''
        current_identification_method = ''

        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=3, max_col=4, values_only=True), start=3
        ):
            serial, violation_name, phenomenon_name, identification_method = map(_cell_text, row)
            if not any((serial, violation_name, phenomenon_name, identification_method)):
                continue

            if bool(serial) != bool(violation_name):
                raise ValueError(f'第 {row_number} 行的序号和违例概念必须同时填写，或同时留空。')

            if serial:
                current_identification_method = ''
                current_violation_id = f'violation:{serial}'
                if current_violation_id not in node_ids:
                    nodes.append(
                        GraphNode(
                            source_id=current_violation_id,
                            uuid=stable_uuid('node', current_violation_id),
                            type='ViolationConcept',
                            name=violation_name,
                            properties={'serial': serial},
                        )
                    )
                    node_ids.add(current_violation_id)
            elif not current_violation_id:
                raise ValueError(f'第 {row_number} 行没有可继承的违例概念。')

            if not phenomenon_name:
                raise ValueError(f'第 {row_number} 行缺少现象。')
            if identification_method:
                current_identification_method = identification_method
            elif not current_identification_method:
                raise ValueError(f'第 {row_number} 行没有可继承的识别方法。')

            # A phenomenon is unique inside one violation. Its description can be
            # changed without generating a duplicate because it is not part of the key.
            phenomenon_id = f'phenomenon:{current_violation_id}:{phenomenon_name}'
            if phenomenon_id in node_ids:
                raise ValueError(f'第 {row_number} 行的现象在同一违例概念下重复：{phenomenon_name!r}。')
            nodes.append(
                GraphNode(
                    source_id=phenomenon_id,
                    uuid=stable_uuid('node', phenomenon_id),
                    type='Phenomenon',
                    name=phenomenon_name,
                    properties={
                        'violation_source_id': current_violation_id,
                        'identification_method': current_identification_method,
                    },
                )
            )
            node_ids.add(phenomenon_id)

            edge_key = f'has_phenomenon:{current_violation_id}:{phenomenon_id}'
            if edge_key not in edge_ids:
                edges.append(
                    GraphEdge(
                        uuid=stable_uuid('edge', edge_key),
                        source_id=current_violation_id,
                        target_id=phenomenon_id,
                        relation='has_phenomenon',
                    )
                )
                edge_ids.add(edge_key)
    finally:
        workbook.close()

    if not nodes:
        raise ValueError('Excel 中没有可导入的数据。')
    return GraphSnapshot(nodes=nodes, edges=edges)
